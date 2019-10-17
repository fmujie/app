import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # cell = sheet['a1']
    # sheet.cell(1, 1)
    # print(cell.value)
    # print(sheet.max_row)
    for row in range(2, sheet.max_row + 1):
        # print(row)
        cell = sheet.cell(row, 2)
        # print(cell.value)
        update_data = cell.value * 0.9
        storage_data = sheet.cell(row, 3)
        storage_data.value = update_data

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=3,
                       max_col=3)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'd2')
    wb.save(filename)

process_workbook('test.xlsx')
# import useful_tool
# import random
# from pathlib import Path
# import ecommerce.shipping
# ecommerce.shipping.calc_shipping()
# from ecommerce.shipping import calc_shipping
# calc_shipping()
# import docx
from math import *
# character_name =  "fmujie"
# character_age = "50"
# year = 152
# res = year - int(character_age)
# print(res)
# is_male = True
# name = "Fmujie cool"
#print(name[0:] + " is shirt")
# print(name.lower())
# print(name)
# print(name.upper())
# print(name)
# print(name.upper().isupper())
# print(len(name))
# print(name[-1])
# print(name.index("F"))
# print(name.index("mu"))
# print(name)
# print(name.replace("shirft", "cool"))
# print(name)
# print(2.3231 + 4)
# my_num = -6
# print(str(my_num) + " is my favorite number")
# print(abs(my_num))
# print(pow(4, 5))
# print(max(4, 6))
# print(min(4, 6))
# print(round(3.2))
# print(round(3.5))
# print(floor(3.7))##抓取不大于
# print(ceil(3.4))
# print(sqrt(36))
# username = input("Enter your name: ")
# print("Hello " + username)
# num1 = input("Enter num1: ")
# method = print("Enter your method")
# num2 = input("Enter num2: ")
# result = int(num1) + int(num2)
# result = float(num1) + float(num2)
# print(result)
# friends = ["hanwei", 2, False]
# friends = ["hanwei", "zhang", "weiwei", "chao", "long"]
# print(friends[-2])
# name_index = friends.index("weiwei")
# print(name_index)
# friends[name_index] = "hah"
# print(friends[name_index])
# print(friends)
# print(friends[1:3])
# new_friends = ["hanwei", "zhang", "weiwei", "chao", "long"]
# luck_numbers = [3, 6, 8, 9, 7]
# print(new_friends)
# print(luck_numbers)
# new_friends.append("wo")
# new_friends.extend(luck_numbers)
# new_friends.insert(1, "Kelly")
# new_friends.remove("long")
# new_friends.clear()
# new_friends.pop()
# print(new_friends.count("weiwei"))
# print(new_friends)
# print(luck_numbers)
# luck_numbers.sort()
# print(luck_numbers)
# luck_numbers.reverse()
# print(luck_numbers)
# new_friends2 = new_friends
# new_friends2 = new_friends.copy()
# print(new_friends2)

# def cube(num):
#     return num*num*num
# result = cube(5)
# print(result)

# is_male = True
# is_tall = True

# if is_male:
#     print("you are male")
# else:
#     print("you are not a male")

# if is_male or is_tall: # and
#     print("you are male or tall")
# elif is_male and not(is_tall):
#     print("somthing")#not() -> !
# else:
#     print("you are not a male nor tall")

# def max_num(num1, num2, num3):
#     if num1 >= num2 and num1>=num3:
#         return num1
#     elif num2 >= num1 and num2 >=num3:
#         return num2
#     else:
#         return num3
#
# print(max_num(3, 4, 5))

# num1 = float(input("Enter your first number: "))
# op = input("Enter operator: ")
# num2 = float(input("Enter your second number: "))
#
# if op == "+":
#     print(num1 + num2)
# elif op == "-":
#     print(num1 - num2)
# elif op == "*":
#     print(num1 * num2)
# elif op == "/":
#     print(num1 / num2)
# else:
#     print("Invalid operator")

# monthConversions = {
#     "Jan": "january",
#     "Feb": "February",
#     2: "April",
# }
#
# print(monthConversions["Feb"])
# print(monthConversions.get("Jan"))
# print(monthConversions.get("Dec", "Not a valid Key"))

# i = 1
# while i <= 10:
#     print(i)
#     i += 1

# secret_word = "giraffe"
# guess = ""
# guess_count = 0
#
# while guess != secret_word:
#     if guess_count >= 3:
#         print("You are loss")
#         break
#     else:
#         guess = input("Enter guess: ")
#         guess_count += 1
# if guess_count <= 3:
#     print("You are win")

# secret_word = "giraffe"
# guess = ""
# guess_count = 0
# guess_limit = 3
# out_of_guesses = False
#
# while guess != secret_word and not(out_of_guesses):
#     if guess_count < guess_limit:
#         guess = input("Enter guess: ")
#         guess_count += 1
#     else:
#         out_of_guesses = True
# if out_of_guesses:
#     print("Out of Guesses, you Lose")
# else:
#     print("You are Win")

# for letter in "fmujie":
#     print(letter)

friends = ["fmu", "mujie"]
# for friend in friends:
#     print(friend)

# for index in range(10):#rang(3, 10)
#     print(index)
#
# for index in range(len(friends)):
#     print(friends[index])

# def raise_to_pow(base_num, pow_num):
#     result = 1
#     for index in range(pow_num):
#         result = result * base_num
#     return result
# print(raise_to_pow(2, 3))

# number_grid = [
#     [1, 2, 3],
#     [4, 5, 6],
#     [0]
# ]
#
# print(number_grid[0][0])
#
# for row in number_grid:
#     print(row)
#
# for row in number_grid:
#     for col in row:
#         print(col)

# def translate(phrase):
#     translation = ""
#     for letter in phrase:
#         if letter.lower() in "aeiou":
#             if letter.isupper():
#                 translation = translation + "G"
#             else:
#                 translation = translation + "g"
#         else:
#             translation = translation + letter
#     return translation
#
# print(translate(input("Enter a phrase: ")))

'''
dsgsgsdv
dfsdgsd
ds
'''

# try:
#     result = 10 / 0
#     number = int(input("Enter anumber: "))
#     print(number)
# except ZeroDivisionError as err:
#     print(err)
# except ValueError:
#     print("Invalid Input")


# test_file = open("test.txt", "r")
#
# able = test_file.readable()
# if able:
#     for each_line in test_file.readlines():
#         print(each_line)
# else:
#     print("打开文件失败")
#
# test_file.close()

# test_file = open("test1.txt", "a")
#
# able = test_file.writable()
# if able:
#     test_file.write("\nwoxihuan")
# else:
#     print("打开文件失败")
#
# test_file.close()

# print(useful_tool.roll_dice(10))

# numbers = [2, 2, 4, 6, 3, 4, 6, 1]
# uniques = []
# for number in numbers:
#     if number not in uniques:
#         uniques.append(number)
# print(uniques)

# coordinates = (1, 2, 3) #[]

# x = coordinates[0]
# y = coordinates[1]
# z = coordinates[2]

# x, y, z = coordinates
# print(x)

# def greet_user(name):
#     print(f"hi {name}")
# greet_user("fmujie")

# class Point:
#     def __init__(self, x, y):
#         self.x = x
#         self.y = y
#     def move(self):
#         print("move")
#     def draw(self):
#         print("draw")

# point1 = Point()
# point1.draw()
# point1.x = 1
# point = Point(10, 20)
# print(point.x)

# class Animal:
#     def walk(self):
#         print("walk")
#
#
# class Dog(Animal):
#     def bark(self):
#         print("bark")
#
#
# class Cat(Animal):
#     pass
#
# dog1 = Dog()
# dog1.bark()

# path = Path()
# print(path.exists())
# print(path.mkdir())
# print(path.rmdir())
# for file in path.glob('*'):
#     print(file)


